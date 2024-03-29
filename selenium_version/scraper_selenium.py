import time
import queue
from threading import Event, Thread, Lock
from selenium_version.searchselenium import SeleniumSearch
from selenium_version.chapters_scraper import ChapterScraper
from openpyxl import Workbook

class YoutubeChapterScraperSelenium:
    def __init__(self, url, excelname, limit, update_ui_queue: queue.Queue, event: Event) -> None:
        self.url = url
        self.excelname: str = excelname
        self.limit = limit
        self.update_ui_queue = update_ui_queue
        self.event = event
        self.video_ids = []
        self.chapters = {}
        self.selenium_search = SeleniumSearch(self)
        self.chapter_scraper1 = ChapterScraper(self)
        self.chapter_scraper2 = ChapterScraper(self)
        self.chapter_scraper3 = ChapterScraper(self)
        self.video_search_done = False
        self.video_processed = 0
        self.total_chapters_found = 0
        self.completed = set()
        self.thread1 = None
        self.thread2 = None
        self.thread3 = None

    def startScraping(self):
        # start chapter scraper loop that will pool the video ids and keep updating the self.chapters
        self.start_chapter_scraper_thread()
        self.selenium_search.search_results(self.url, self.limit) # this will keep filling video ids
        self.video_search_done = True
        self.thread1.join()
        self.thread2.join()
        self.thread3.join()

    def start_chapter_scraper_thread(self):
        self.thread1 = Thread(target=self.chapter_scraper1.scrape_chapters)
        self.thread2 = Thread(target=self.chapter_scraper2.scrape_chapters)
        self.thread3 = Thread(target=self.chapter_scraper3.scrape_chapters)
        self.thread1.start()
        self.thread2.start()
        self.thread3.start()

    def save_data(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Youtube Chapter Name"
        ws['B1'] = "Number of Occurrences"
        
        # Sort chapters by number of occurrences
        sorted_chapters = sorted(self.chapters.items(), key=lambda x: x[1], reverse=True)
        
        row = 2 
        for chapter_name, occurrences in sorted_chapters:
            ws[f'A{row}'] = chapter_name
            ws[f'B{row}'] = occurrences
            row += 1
        
        filename = self.excelname.replace(" ", "_")

        if '.xlsx' not in filename:
            filename += '.xlsx'
        wb.save(f"data/{filename}")

    
    def update_chapters(self, chapters: str):
        for chapter in chapters:
            ch = chapter.title()
            if ch in self.chapters:
                self.chapters[ch] += 1
            else:
                self.chapters[ch] = 1
    