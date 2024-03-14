import time
import queue
from threading import Event, Thread, Lock
from search_videos.searchselenium import SeleniumSearch
from youtube_scraper.chapters_scraper import ChapterScraper
from openpyxl import Workbook

class YoutubeChapterScraperSelenium:
    def __init__(self, url, excelname, limit, update_ui_queue: queue.Queue, event: Event) -> None:
        self.url = url
        self.excelname = excelname
        self.limit = limit
        self.update_ui_queue = update_ui_queue
        self.event = event
        self.video_ids = []
        self.chapters = {}
        self.selenium_search = SeleniumSearch(self)
        self.chapter_scraper = ChapterScraper(self)
        self.video_search_done = False
        self.thread = None

    def startScraping(self):
        # start chapter scraper loop that will pool the video ids and keep updating the self.chapters
        self.start_chapter_scraper_thread()
        self.selenium_search.search_results(self.url, self.limit) # this will keep filling video ids
        self.video_search_done = True
        self.thread.join()

    def start_chapter_scraper_thread(self):
        self.thread = Thread(target=self.chapter_scraper.scrape_chapters)
        self.thread.start()

    def save_data(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Youtube Chapter Name"
        ws['B1'] = "Number of Occurrences"
        row = 2 
        for chapter_name, occurrences in self.chapters.items():
            ws[f'A{row}'] = chapter_name
            ws[f'B{row}'] = occurrences
            row += 1
        filename = self.excelname
        if '.xlsx' not in filename:
            filename += '.xlsx'
        wb.save(f"data/{filename}")
    
    def update_chapters(self, chapters: str):
        for chapter in chapters:
            ch = chapter.title()
            if ch in self.chapters:
                self.chapters[ch] += 1
            else:
                self.chapters[ch] = 1
    