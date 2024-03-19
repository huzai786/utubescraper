import time
import queue
from threading import Event, Thread
from api_version.apisearch import ApiSearch
from api_version.chapters_scraper import ChapterScraper
from openpyxl import Workbook
from settings import api_key

class YoutubeChapterScraperApi:
    def __init__(self, query, excelname, limit, update_ui_queue: queue.Queue, event: Event, published_after=None, published_before=None) -> None:
        self.query = query
        self.excelname: str = excelname
        self.limit = limit
        self.update_ui_queue = update_ui_queue
        self.event = event
        self.video_ids = []
        self.chapters = {}
        self.api_search = ApiSearch(api_key, self)
        self.chapter_scraper1 = ChapterScraper(self)
        self.chapter_scraper2 = ChapterScraper(self)
        self.chapter_scraper3 = ChapterScraper(self)
        self.video_processed = 0
        self.total_chapters_found = 0
        self.completed = set()
        self.thread1 = None
        self.thread2 = None
        self.thread3 = None
        self.published_after = published_after
        self.published_before = published_before

    def startScraping(self):
        self.video_ids = self.api_search.search_results(self.query, self.limit, self.published_after, self.published_before)
        self.start_chapter_scraper_thread()


    def start_chapter_scraper_thread(self):
        self.thread1 = Thread(target=self.chapter_scraper1.scrape_chapters)
        self.thread2 = Thread(target=self.chapter_scraper2.scrape_chapters)
        self.thread3 = Thread(target=self.chapter_scraper3.scrape_chapters)
        self.thread1.start()
        self.thread2.start()
        self.thread3.start()
        self.thread1.join()
        self.thread2.join()
        self.thread3.join()

    def save_data(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Youtube Chapter Name"
        ws['B1'] = "Number of Occurrences"
        sorted_chapters = sorted(self.chapters.items(), key=lambda x: x[1], reverse=True)
        row = 2 
        for chapter_name, occurrences in sorted_chapters:
            ws[f'A{row}'] = chapter_name
            ws[f'B{row}'] = occurrences
            row += 1
        
        filename = self.excelname.replace(" ", "_")

        if '.xlsx' not in filename:
            filename += '.xlsx'
        wb.save(f"data/{filename}")

    
    def update_chapters(self, chapters: str):
        for chapter in chapters:
            ch = chapter.title()
            if ch in self.chapters:
                self.chapters[ch] += 1
            else:
                self.chapters[ch] = 1
    